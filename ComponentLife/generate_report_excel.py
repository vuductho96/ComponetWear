# generate_report_excel.py — Native Dynamic OpenXML Excel Report Generator for ComponentWear
import sys
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

def format_report_period_title(period_label):
    if not period_label or period_label == 'All':
        return 'TẤT CẢ THỜI GIAN'
    import re
    m = re.match(r'^([A-Za-z]{3})-(\d{4})$', str(period_label))
    if m:
        months = {'Jan':'01', 'Feb':'02', 'Mar':'03', 'Apr':'04', 'May':'05', 'Jun':'06', 'Jul':'07', 'Aug':'08', 'Sep':'09', 'Oct':'10', 'Nov':'11', 'Dec':'12'}
        return f"THÁNG {months.get(m.group(1), m.group(1))}/{m.group(2)}"
    m2 = re.match(r'^(\d{4})-(\d{2})$', str(period_label))
    if m2:
        return f"THÁNG {m2.group(2)}/{m2.group(1)}"
    if re.match(r'^\d{4}$', str(period_label)):
        return f"NĂM {period_label}"
    return str(period_label).upper()

def build_excel_report(data_json_path, output_excel_path):
    with open(data_json_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)

    period_label = payload.get('periodLabel', 'All Historical Data')
    mold_filter = payload.get('moldFilter', 'All')
    generated_at = payload.get('generatedAt', '')
    top10 = payload.get('top10', [])
    components = payload.get('components', [])
    replacements = payload.get('replacements', [])

    wb = openpyxl.Workbook()

    # Define color palette & styles
    navy_dark = "1E293B"
    navy_light = "334155"
    gray_bg = "F8FAFC"
    border_gray = "E2E8F0"

    font_title = Font(name="Segoe UI", size=14, bold=True, color=navy_dark)
    font_sub = Font(name="Segoe UI", size=9.5, italic=False, color="64748B")
    font_sec_header = Font(name="Segoe UI", size=11, bold=True, color=navy_dark)
    font_tbl_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=9, color="0F172A")
    font_data_bold = Font(name="Segoe UI", size=9, bold=True, color="0F172A")

    fill_tbl_header = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    fill_zebra = PatternFill(start_color=gray_bg, end_color=gray_bg, fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=border_gray),
        right=Side(style='thin', color=border_gray),
        top=Side(style='thin', color=border_gray),
        bottom=Side(style='thin', color=border_gray)
    )
    header_border = Border(
        left=Side(style='thin', color=navy_light),
        right=Side(style='thin', color=navy_light),
        top=Side(style='thin', color=navy_dark),
        bottom=Side(style='thin', color=navy_dark)
    )

    # ==========================================
    # SHEET 1: 01_REPORT (Native Dynamic Chart)
    # ==========================================
    ws1 = wb.active
    ws1.title = "01_REPORT"
    ws1.views.sheetView[0].showGridLines = False

    # Page Setup (A4 Landscape, 1-page fit)
    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
    ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 1
    ws1.page_setup.fitToPage = True

    # Column Widths
    col_widths = {
        'A': 4, 'B': 8, 'C': 16, 'D': 18, 'E': 14, 'F': 14,
        'G': 16, 'H': 15, 'I': 13, 'J': 13, 'K': 15, 'L': 16, 'M': 4
    }
    for col, w in col_widths.items():
        ws1.column_dimensions[col].width = w

    formatted_period = format_report_period_title(period_label)

    # Header
    ws1['B2'] = f"TOP 10 LINH KIỆN THAY NHIỀU NHẤT - {formatted_period}"
    ws1['B2'].font = font_title
    ws1['B2'].alignment = Alignment(vertical='center')
    ws1.row_dimensions[2].height = 24

    # Set row heights for chart placement (Rows 4 to 20)
    for r in range(4, 21):
        ws1.row_dimensions[r].height = 18

    # Section Table Header (Row 22)
    ws1['B22'] = f"BẢNG CHI TIẾT TOP 10 LINH KIỆN THAY NHIỀU NHẤT ({formatted_period})"
    ws1['B22'].font = font_sec_header
    ws1['B22'].alignment = Alignment(vertical='center')
    ws1.row_dimensions[22].height = 20

    # Table Column Headers (Row 23)
    headers1 = [
        ("B", "Top", "center"),
        ("C", "Mã Linh Kiện", "left"),
        ("D", "Tên Khuôn", "left"),
        ("E", "Series", "center"),
        ("F", "Lượt Thay", "right"),
        ("G", "Số Lượng (Pcs)", "right"),
        ("H", "TB Shot / CK", "right"),
        ("I", "Min Shot", "right"),
        ("J", "Max Shot", "right"),
        ("K", "Shot HT", "right"),
        ("L", "Lần Thay Cuối", "center")
    ]

    ws1.row_dimensions[23].height = 20
    for col_letter, label, align in headers1:
        cell = ws1[f"{col_letter}23"]
        cell.value = label
        cell.font = font_tbl_header
        cell.fill = fill_tbl_header
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = header_border

    # Populate Top 10 Table Rows (Rows 24 to 33)
    start_row = 24
    if not top10:
        ws1.merge_cells(f"B{start_row}:L{start_row}")
        cell = ws1[f"B{start_row}"]
        cell.value = "Không có dữ liệu thay thế trong kỳ lọc được chọn."
        cell.font = Font(name="Segoe UI", size=9.5, italic=True, color="64748B")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[start_row].height = 22
        end_row = start_row
    else:
        for idx, r in enumerate(top10):
            row_num = start_row + idx
            ws1.row_dimensions[row_num].height = 18
            fill_current = fill_zebra if idx % 2 == 1 else fill_white

            row_data = [
                ("B", idx + 1, "center", False, "0"),
                ("C", str(r.get('part', '')), "left", True, None),
                ("D", str(r.get('moldName', '')), "left", False, None),
                ("E", str(r.get('series', '-')), "center", False, None),
                ("F", int(r.get('replacementCount', 0)), "right", False, "#,##0"),
                ("G", int(r.get('totalPartsReplacedPcs', 0)), "right", True, "#,##0"),
                ("H", int(r['averageShotLife']) if r.get('averageShotLife') is not None else "N/A", "right", False, "#,##0" if r.get('averageShotLife') is not None else None),
                ("I", int(r['minShotLife']) if r.get('minShotLife') is not None else "N/A", "right", False, "#,##0" if r.get('minShotLife') is not None else None),
                ("J", int(r['maxShotLife']) if r.get('maxShotLife') is not None else "N/A", "right", False, "#,##0" if r.get('maxShotLife') is not None else None),
                ("K", int(r['currentShotCount']) if r.get('currentShotCount') is not None else "N/A", "right", False, "#,##0" if r.get('currentShotCount') is not None else None),
                ("L", str(r.get('lastReplacementDate', 'N/A') or 'N/A'), "center", False, None)
            ]

            for col_letter, val, align, bold, num_fmt in row_data:
                cell = ws1[f"{col_letter}{row_num}"]
                cell.value = val
                cell.font = font_data_bold if bold else font_data
                cell.fill = fill_current
                cell.alignment = Alignment(horizontal=align, vertical='center')
                cell.border = thin_border
                if num_fmt and isinstance(val, (int, float)):
                    cell.number_format = num_fmt

        end_row = start_row + len(top10) - 1

        # ========================================================
        # NATIVE EXCEL COMBO CHART (Columns = Lượt thay, Line = Pcs)
        # ========================================================
        from openpyxl.chart.label import DataLabelList

        # 1. Bar Chart for Replacement Count (Col F)
        chart_bar = BarChart()
        chart_bar.title = f"TOP 10 LINH KIỆN THAY NHIỀU NHẤT - {formatted_period}"
        chart_bar.style = 10
        chart_bar.y_axis.title = "Số lượt thay"
        chart_bar.x_axis.title = "Mã Linh Kiện"
        chart_bar.width = 25
        chart_bar.height = 11

        data_bar = Reference(ws1, min_col=6, min_row=23, max_row=end_row) # Col F: Lượt Thay
        categories = Reference(ws1, min_col=3, min_row=24, max_row=end_row) # Col C: Mã Linh Kiện
        chart_bar.add_data(data_bar, titles_from_data=True)
        chart_bar.set_categories(categories)
        chart_bar.y_axis.majorGridlines = None

        # 2. Line Chart for Replacement Quantity (Col G) on secondary axis
        chart_line = LineChart()
        data_line = Reference(ws1, min_col=7, min_row=23, max_row=end_row) # Col G: Số Lượng (Pcs)
        chart_line.add_data(data_line, titles_from_data=True)
        chart_line.y_axis.title = "Số lượng (Pcs)"
        chart_line.y_axis.axId = 200
        chart_line.y_axis.crosses = "max"
        chart_line.dataLabels = DataLabelList()
        chart_line.dataLabels.showVal = True
        chart_line.y_axis.majorGridlines = None

        # Combine into Combo Chart
        chart_bar += chart_line
        if chart_bar.legend:
            chart_bar.legend.legendPos = "b"

        # Place chart at B4
        ws1.add_chart(chart_bar, "B4")

    # ==========================================
    # SHEET 2: 02_COMPONENT_SUMMARY
    # ==========================================
    ws2 = wb.create_sheet(title="02_COMPONENT_SUMMARY")
    ws2.views.sheetView[0].showGridLines = True
    ws2.freeze_panes = "A2"

    summary_headers = [
        "No.", "Mã Linh Kiện (Part)", "Khuôn (Series/Old/New)",
        "Số Lượt Thay", "Số Lượng (Pcs)", "Số Chu Kỳ", "TB Shot / CK",
        "Min Shot", "Max Shot", "Shot HT", "Tổng Shot Tích Lũy", "Lần Thay Cuối"
    ]
    ws2.append(summary_headers)
    ws2.row_dimensions[1].height = 22

    for col_idx, h in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_tbl_header
        cell.fill = fill_tbl_header
        cell.alignment = Alignment(horizontal='right' if col_idx in [4,5,6,7,8,9,10,11] else 'center', vertical='center')

    replaced_components = [c for c in components if int(c.get('replacementCount', 0)) > 0 or int(c.get('totalPartsReplacedPcs', 0)) > 0]

    for idx, c in enumerate(replaced_components):
        row_vals = [
            idx + 1,
            c.get('part', ''),
            c.get('moldSeries', '') or f"{c.get('series', '-')}/{c.get('moldName', '')}",
            int(c.get('replacementCount', 0)),
            int(c.get('totalPartsReplacedPcs', 0)),
            int(c.get('cycleCount', 0)),
            int(c['averageShotLife']) if c.get('averageShotLife') is not None else "N/A",
            int(c['minShotLife']) if c.get('minShotLife') is not None else "N/A",
            int(c['maxShotLife']) if c.get('maxShotLife') is not None else "N/A",
            int(c['currentShotCount']) if c.get('currentShotCount') is not None else "N/A",
            int(c.get('totalLifetimeShots', 0)),
            c.get('lastReplacementDate', 'N/A') or 'N/A'
        ]
        ws2.append(row_vals)
        r_num = idx + 2
        ws2.row_dimensions[r_num].height = 18
        for col_idx in range(1, len(summary_headers) + 1):
            cell = ws2.cell(row=r_num, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if col_idx in [4,5,6,7,8,9,10,11] and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_idx in [1, 12]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    ws2.auto_filter.ref = f"A1:L{len(replaced_components) + 1}"

    # Auto column widths for Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 13)

    # ==========================================
    # SHEET 3: 03_REPLACEMENT_LOG
    # ==========================================
    ws3 = wb.create_sheet(title="03_REPLACEMENT_LOG")
    ws3.views.sheetView[0].showGridLines = True
    ws3.freeze_panes = "A2"

    log_headers = [
        "No.", "Ngày Thay", "Mã Linh Kiện (Part)", "Khuôn (Series/Old/New)",
        "Số Lượng (Pcs)", "Mã Yêu Cầu (Req ID)", "Ghi Chú / Serial"
    ]
    ws3.append(log_headers)
    ws3.row_dimensions[1].height = 22

    for col_idx, h in enumerate(log_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = font_tbl_header
        cell.fill = fill_tbl_header
        cell.alignment = Alignment(horizontal='right' if col_idx == 5 else 'center', vertical='center')

    for idx, r in enumerate(replacements):
        qty = int(r.get('Label', 1)) if str(r.get('Label', '')).isdigit() else 1
        mold_s = r.get('moldSeries', '') or f"{r.get('Series', '')}/{r.get('DieSet', '')}"
        row_vals = [
            idx + 1,
            r.get('ReplaceDate', 'N/A') or 'N/A',
            r.get('Part', ''),
            mold_s,
            qty,
            r.get('RequestId', '') or '',
            r.get('Label', '') or ''
        ]
        ws3.append(row_vals)
        r_num = idx + 2
        ws3.row_dimensions[r_num].height = 18
        for col_idx in range(1, len(log_headers) + 1):
            cell = ws3.cell(row=r_num, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_idx in [1, 2, 6]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    ws3.auto_filter.ref = f"A1:G{len(replacements) + 1}"

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_excel_path)
    print(f"REPORT SAVED: {output_excel_path}")

if __name__ == '__main__':
    if len(sys.argv) >= 3:
        build_excel_report(sys.argv[1], sys.argv[2])
    else:
        print("Usage: python generate_report_excel.py <data.json> <output.xlsx>")
